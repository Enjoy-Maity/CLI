import logging
import os
import traceback
import numpy as np
import pandas as pd
from tkinter import messagebox
from pathlib import Path
from Custom_Exception import CustomException
from threading import Thread
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Side,Border,PatternFill,Alignment
from openpyxl.styles.colors import Color

# log_file_path = "C:/Ericsson_Application_Logs/CLI_Automation_Logs/"
# Path(log_file_path).mkdir(parents=True,exist_ok=True)
# log_file = os.path.join(log_file_path,"Sheet_Creation_Task.log")
# logging.basicConfig(filename=log_file,
#                              filemode="a",
#                              format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: ({'%(module)s'}): {'%(message)s'}",
#                              datefmt='%d-%b-%Y %I:%M:%S %p',
#                              encoding= "UTF-8",
#                              level=logging.DEBUG)
# logging.captureWarnings(capture=True)

# flag = ''

def sheet_creater(**kwargs) -> str:
    logging.basicConfig(filename=log_file,
                                filemode="a",
                                format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: ({'%(module)s'}): {'%(message)s'}",
                                datefmt='%d-%b-%Y %I:%M:%S %p',
                                encoding= "UTF-8",
                                level=logging.DEBUG)
    logging.captureWarnings(capture=True)
    
    try:
        logging.info(f"Starting the Sheet Creater for {os.path.basename(kwargs['file'])}")
        host_ips_sheets_required = kwargs['host_ips_sheets_required']
        file                     = kwargs['file']
        standard_input           = kwargs['standard_design_template_path']
        file_name                = os.path.basename(file)
        standard_input_file_name = os.path.basename(standard_input)

        logging.debug(f"Reading the '{standard_input_file_name}' for the 'Standard Template' worksheet using 'openpyxl'")
        standard_input_workbook = load_workbook(standard_input)
        standard_input_template_worksheet = standard_input_workbook['Standard Template']
        max_rows                          = standard_input_template_worksheet.max_row
        max_columns                       = standard_input_template_worksheet.max_column

        logging.debug(f"Reading the {file_name} using openpyxl")
        host_ips_sheets_required_workbook = load_workbook(file)
        host_ips_sheetnames_present_in_the_workbook = host_ips_sheets_required_workbook.sheetnames

        logging.debug("Getting the length for widths for the columns of the worksheet")
        width_list = []
        i = 1
        while(i <= max_rows):
            j = 1
            while(j <= max_columns):
                cell_selected = f'{get_column_letter(j)}{i}'
                if(len(width_list) < j):
                    width_list.insert(j-1,len(str(standard_input_template_worksheet[cell_selected].value))+3)

                else:
                    if((len(str(standard_input_template_worksheet[cell_selected].value))+3) > width_list[j-1]):
                        width_list[j-1] = len(str(standard_input_template_worksheet[cell_selected].value))+3
                j+=1
            i+=1

        # print(host_ips_sheetnames_present_in_the_workbook)

        i = 0
        while(i < host_ips_sheets_required.size):
            logging.debug(f"Checking for presence of {host_ips_sheets_required[i]} in {file_name} \n")
            logging.debug(f"Checking the condition for {not host_ips_sheets_required[i] in host_ips_sheetnames_present_in_the_workbook}")
            if(not host_ips_sheets_required[i] in host_ips_sheetnames_present_in_the_workbook):

                logging.debug(f"Creating {host_ips_sheets_required[i]} sheet in {file_name}")
                worksheet = host_ips_sheets_required_workbook.create_sheet(title=host_ips_sheets_required[i])

                j = 1
                while(j <= max_rows):
                    k = 1
                    while(k <= max_columns):
                        cell_selected = f'{get_column_letter(k)}{j}'
                        # color = standard_input_template_worksheet[cell_selected].fill.start_color
                        side = Side(color=Color(rgb='00000000'), style='thin')
                        # logging.debug(f"Writing {cell_selected} for {host_ips_sheets_required[i]} in {file_name}")

                        worksheet[cell_selected].value = standard_input_template_worksheet[cell_selected].value
                        # with open(r"C:\Users\emaienj\Downloads\VPLS_CLI_Design_Documents\VPLS_CLI_Design_Documents\demo.txt","a+") as f:
                        #     f.write(f"color of {get_column_letter(k)}{j} =====> {color}\n\n")
                        # if(not ( (color.rgb == None) or (color == Color(rgb = '00000000')) ) ):
                        #     # font.bold = True
                        #     worksheet[cell_selected].fill = PatternFill(start_color = color,
                        #                                                 end_color = color,
                        #                                                 bgColor = color,
                        #                                                 fill_type = 'solid')

                        # worksheet[cell_selected].font = 
                        # worksheet[cell_selected].alignment = Alignment(horizontal='center',
                        #                                                vertical='center')
                        
                        worksheet[cell_selected].font      = standard_input_template_worksheet[cell_selected].font.copy()
                        worksheet[cell_selected].alignment = standard_input_template_worksheet[cell_selected].alignment.copy()
                        worksheet[cell_selected].fill      = standard_input_template_worksheet[cell_selected].fill.copy()
                        worksheet[cell_selected].border    = Border(left = side,
                                                                    right = side,
                                                                    top = side,
                                                                    bottom = side)

                        k+=1
                    j+=1

                j = 1
                while(j <= max_columns):
                    worksheet.column_dimensions[get_column_letter(j)].width = width_list[j-1]
                    j+=1
            i+=1

        logging.info("Removing the extra sheet that are not required in the workbook")
        for sheet in host_ips_sheetnames_present_in_the_workbook:
            if(not (sheet in host_ips_sheets_required)):
                del host_ips_sheets_required_workbook[sheet]

        logging.info(f"Saving the file {file_name}")
        host_ips_sheets_required_workbook.save(file)

        host_ips_sheets_required_workbook.close()
        del host_ips_sheets_required_workbook

        logging.debug(f"Closing the {standard_input_file_name}")
        standard_input_workbook.close()
        del standard_input_workbook
    
    except CustomException:
        global flag;
        flag = 'Unsuccessful'
        logging.error(f"{traceback.format_exc()}\n\nraised CustomException==>\ntiltle = {e.title}\nmessage = {e.message}")

    except Exception as e:
        flag = 'Unsuccessful'
        logging.error(f"{traceback.format_exc()}\n\nException:==>{e}")
        messagebox.showerror("Exception Occurred!",e)

    
def file_creater(**kwargs):
    logging.basicConfig(filename=log_file,
                                filemode="a",
                                format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: ({'%(module)s'}): {'%(message)s'}",
                                datefmt='%d-%b-%Y %I:%M:%S %p',
                                encoding= "UTF-8",
                                level=logging.DEBUG)
    logging.captureWarnings(capture=True)
    
    file = kwargs['file']
    wb = Workbook()
    wb.save(file)
    wb.close()
    del wb

# kwargs.keys() ==> (file_name)
def main_func(**kwargs) -> str:
    log_file_path = "C:/Ericsson_Application_Logs/CLI_Automation_Logs/"
    Path(log_file_path).mkdir(parents=True,exist_ok=True)
    global log_file; log_file = os.path.join(log_file_path,"Sheet_Creation_Task.log")
    logging.basicConfig(filename=log_file,
                                filemode="a",
                                format=f"[ {'%(asctime)s'} ]: <<{'%(levelname)s'}>>: ({'%(module)s'}): {'%(message)s'}",
                                datefmt='%d-%b-%Y %I:%M:%S %p',
                                encoding= "UTF-8",
                                level=logging.DEBUG)
    logging.captureWarnings(capture=True)

    global flag; flag = ''

    logging.info("#############################################<<Starting the Sheet Creation Task>>#################################################")
    
    try:
        host_details_file_name = kwargs['file_name']
        file_reader = pd.ExcelFile(host_details_file_name,engine='openpyxl')
        host_details_file_sheetnames = file_reader.sheet_names

        # Checking for the sheet 'Host Details' in the uploaded workbook.
        logging.debug("Checking for the 'Host Details' worksheet in the uploaded workbook")
        if(not "Host Details" in host_details_file_sheetnames):
            raise CustomException("Sheet Not Found!","'Host Details' workbook not found in the uploaded workbook, Kindly check and try again!")
        
        logging.debug("Creating dataframe for the host details sheet dataframe")
        # Creating the host details sheet dataframe
        host_details_df = pd.read_excel(file_reader,sheet_name='Host Details')
        
        logging.info(f'Read the host details\n\n{host_details_df}')

        # Performing all the mandatory checks for the 'Host Details' sheet.
        unique_host_ips = host_details_df['Host_IP'].unique()

        # if(pd.NA in nan_test ):
        nan_test = any(pd.isna(element) for element in unique_host_ips)
        
        # if(np.nan in nan_test):
        #     nan_test = any(pd.isna(element) for element in unique_host_ips)            
        blank_Sr_no_list = []
        # If nan_test is true
        if(nan_test):
            i = 0
            while(i<len(host_details_df)):
                if(pd.isna(host_details_df.iloc[i]['Host_IP'])):
                    blank_Sr_no_list.append(host_details_df.iloc[i]['Sr.No'])
                i+=1
        
        # Finding row with blank Vendor
        blank_Sr_no_list_for_Vendor = []
        unique_vendors_in_host_details = host_details_df['Vendor'].unique()
        
        logging.info("Checking for rows with blank Vendor details")
        nan_test = any(pd.isna(element) for element in unique_vendors_in_host_details)
        
        if(nan_test):
            i =0
            while(i<len(host_details_df)):
                if(pd.isna(host_details_df.iloc[i]['Vendor'])):
                    blank_Sr_no_list_for_Vendor.append(host_details_df.iloc[i]['Sr.No'])
                i+=1
        # temp_array = unique_host_ips[~np.isnan(unique_host_ips)]
        # temp_array = unique_host_ips[~pd.isna(unique_host_ips)]
        host_details_df.dropna(inplace= True)
        

        # Checking duplicated Host IPs in the Host Details sheet
        duplicate_host_ip_boolean_series = host_details_df.duplicated(subset=['Host_IP'])
        duplicated_host_ip_Sr_no_list = []
        logging.info("Checking duplicated host ip details in the Host Details sheet")
        
        i = 0
        while(i < duplicate_host_ip_boolean_series.size):
            if(duplicate_host_ip_boolean_series[i]):
                temp_array_list = host_details_df[host_details_df['Host_IP'] == host_details_df.iloc[i]['Host_IP']]
                duplicated_host_ip_Sr_no_list.extend(list(temp_array_list['Sr.No']))
            i+=1
        
        
        if(len(duplicated_host_ip_Sr_no_list) >0):
            duplicated_host_ip_Sr_no_list = np.unique(np.array(duplicated_host_ip_Sr_no_list))
        
        if((len(blank_Sr_no_list)> 0) and (len(duplicated_host_ip_Sr_no_list) > 0) and (len(blank_Sr_no_list_for_Vendor) > 0)):
            file_reader.close()
            del file_reader
            logging.error("Blank Host IP & Vendors and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                    f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}\n\nand \n\nBlank Vendor Details found for Sr no: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}\n\nand \n\nDuplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")
        
        if((len(blank_Sr_no_list)> 0) and (len(duplicated_host_ip_Sr_no_list) > 0)):
            file_reader.close()
            del file_reader
            logging.error("Blank Host IP and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                    f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}\n\nand \n\nDuplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")
        
        if((len(blank_Sr_no_list)> 0) and (len(blank_Sr_no_list_for_Vendor) > 0)):
            file_reader.close()
            del file_reader
            logging.error("Blank Host IP and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                    f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}\n\nand \n\nBlank Vendor Details found for Sr no: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}")
        
        if((len(blank_Sr_no_list_for_Vendor)> 0) and (len(duplicated_host_ip_Sr_no_list) > 0)):
            file_reader.close()
            del file_reader
            logging.error("Blank Vendor and duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                    f"Blank Vendor Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}\n\nand \n\nDuplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")
        
        if(len(blank_Sr_no_list) > 0):
            file_reader.close()
            del file_reader
            logging.error("Blank Host IP details are found!")
            raise CustomException("Blank Host IP Details Found!",
                                    f"Blank IP Details found for Sr no.: {', '.join(str(element) for element in blank_Sr_no_list)}")
        
        if(len(duplicated_host_ip_Sr_no_list) > 0):
            file_reader.close()
            del file_reader
            logging.error("Duplicated Host IP details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                    f"Duplicate Host IPs found for Sr no: {', '.join(str(element) for element in duplicated_host_ip_Sr_no_list)}")
        
        if(len(blank_Sr_no_list_for_Vendor) > 0):
            file_reader.close()
            del file_reader
            logging.error("Blank Vendor details are found!")
            raise CustomException("Host IP Details Incorrect!",
                                    f"Duplicate Host IPs found for Sr no: {', '.join(str(element) for element in blank_Sr_no_list_for_Vendor)}")
        
        # getting the list of all the files present in the parent directory of host_details_file_name
        logging.info("Getting the list of all the files present in the parent directory of host_details_file_name")
        parent_dir = os.path.dirname(host_details_file_name)
        
        list_of_files_present_parent_dir = os.listdir(parent_dir)
        logging.debug(f"List of all the files present in the parent directory of host_details_file\n\n{'\n'.join(str(element) for element in list_of_files_present_parent_dir)}\n\n")

        missing_standard_input_design_template = []
        logging.info("Checking for the availability of Standard Input Design Templates")
        
        i = 0
        while(i < unique_vendors_in_host_details.size):
            if(unique_vendors_in_host_details[i].strip().upper() == 'NOKIA'):
                if(not Path(os.path.join(parent_dir,'Nokia_Design Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Nokia')
            
            if(unique_vendors_in_host_details[i].strip().upper() == 'ERICSSON'):
                if(not Path(os.path.join(parent_dir,'Ericsson_Design Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Ericsson')
            
            if(unique_vendors_in_host_details[i].strip().upper() == 'HUAWEI'):
                if(not Path(os.path.join(parent_dir,'Huawei_Design Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Huawei')
            
            if(unique_vendors_in_host_details[i].strip().upper() == 'CISCO'):
                if(not Path(os.path.join(parent_dir,'Cisco_Design Input_Template.xlsx')).exists()):
                    missing_standard_input_design_template.append('Cisco')
            i+=1
        
        if(len(missing_standard_input_design_template) > 0):
            del host_details_file_name
            file_reader.close()
            del file_reader
            logging.error("Standard Input design template missing")
            raise CustomException("Standard Input Design Template Missing!",
                                  f"Standard Input design template missing for below mentioned Vendors:\n\n{', '.join(missing_standard_input_design_template)}")
        neo_parent_dir = os.path.join(parent_dir,"Design_Input_Sheets")
        Path(neo_parent_dir).mkdir(exist_ok= True, parents=True)
        
        logging.debug(f"Created new folder if not created earlier \'{parent_dir}\'\n")
        
        file_to_be_created = "{}_Design_Input_Sheet.xlsx"
        file_to_be_selected = "{}_Design Input_Template.xlsx"
        
        # Task for creating the host ip sheets in the respective Vendor workbook
        thread_list_for_vendor_sheet_creation = []
        try:
            i = 0
            while(i < unique_vendors_in_host_details.size):
                host_ips_sheets_required = host_details_df[host_details_df['Vendor'] == unique_vendors_in_host_details[i]]
                host_ips_sheets_required = np.array(host_ips_sheets_required['Host_IP'])
                
                # Checking the file existence for the input files for the user
                logging.debug(f"Checking if the '{file_to_be_created.format(unique_vendors_in_host_details[i])}' exists or not, if yes then adding ip sheets, otherwise creating the file itself.")
                logging.debug(f"Checking the condition working or not '{not Path(os.path.join(neo_parent_dir,file_to_be_created.format(unique_vendors_in_host_details[i]))).exists() =}'")

                if(not Path(os.path.join(parent_dir,file_to_be_created.format(unique_vendors_in_host_details[i]))).exists()):
                    file_creater(file = os.path.join(neo_parent_dir,file_to_be_created.format(unique_vendors_in_host_details[i])))

                
                logging.debug(f"Checking for the presence of the 'Standard Template' worksheet in the {file_to_be_selected.format(unique_vendors_in_host_details[i])}")
                temp_reader = pd.ExcelFile(os.path.join(parent_dir,file_to_be_selected.format(unique_vendors_in_host_details[i])))
                
                if(not 'Standard Template' in temp_reader.sheet_names):
                    del temp_reader
                    messagebox.showwarning("Template Sheet Missing!",f"'Standard Template' worksheet missing in {file_to_be_selected.format(unique_vendors_in_host_details[i])}, Kindly Check and Try Again!")
                    continue
                
                logging.info("Deleting the temp_reader (Created to find the existence of 'Standard Template' worksheet)")
                del temp_reader
                
                logging.debug(f"Creating the input file with sheets for '{unique_vendors_in_host_details[i]}' via thread")

                # Creating Sheets for the unique IPs using threads
                thread = Thread(target = sheet_creater,
                                kwargs={'host_ips_sheets_required': host_ips_sheets_required,
                                        'file' :os.path.join(neo_parent_dir,file_to_be_created.format(unique_vendors_in_host_details[i])),
                                        'standard_design_template_path': os.path.join(parent_dir,file_to_be_selected.format(unique_vendors_in_host_details[i]))})
                thread_list_for_vendor_sheet_creation.append(thread)
                thread.daemon = True
                thread.start()
                i+=1
            
            logging.info("Stopping all the threads for vendor workbook after writing the information")
            i = 0
            while(i < len(thread_list_for_vendor_sheet_creation)):
                thread_list_for_vendor_sheet_creation[i].join()
                i+=1
        
        except Exception as e:
            # global flag;
            flag = 'Unsuccessful'
            logging.error(f"{traceback.format_exc()}\n\nException:==>{e}")
            messagebox.showerror("Exception Occurred!",e)

        del host_details_file_name
        file_reader.close()
        del file_reader
        
        if(flag != 'Unsuccessful'):
            logging.info("Setting the flag status to 'Successful'")
            flag = 'Successful'

    except CustomException:
        flag = 'Unsuccessful'
        logging.error(f"{traceback.format_exc()}\n\nraised CustomException==>\ntiltle = {e.title}\nmessage = {e.message}")

    except Exception as e:
        flag = 'Unsuccessful'
        logging.error(f"{traceback.format_exc()}\n\nException:==>{e}")
        messagebox.showerror("Exception Occurred!",e)
    
    finally:
        logging.info(f"Returning Flag\n\n\t{flag=}")
        logging.shutdown()
        return flag
    
# print(main_func(file_name = r"C:\Users\emaienj\Downloads\VPLS_CLI_Design_Documents\VPLS_CLI_Design_Documents\Host_Details.xlsx"))