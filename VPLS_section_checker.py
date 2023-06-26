import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
from openpyxl.utils import get_column_letter
from Custom_Thread import CustomThread
from CustomException import CustomException
from tkinter import messagebox

def dataframe_checker(dataframe,node):
    try:
        for i in range(len(dataframe)):
            if((len(str(dataframe.iloc[i]['VPLS ID']).strip()) == 0 ) or (dataframe.iloc[i]['VPLS ID'] == "TempNA")):
                raise CustomException(" VPLS ID Missing!",f"Kindly check the 'VPLS ID' section in {node} in the selected workbook for missing 'VPLS ID'")
        
        unique_VPLS_ID_len = len(dataframe['VPLS ID'].unique())
        if(unique_VPLS_ID_len < len(dataframe)):
            raise CustomException(" Duplicate VPLS ID Present!",f"Kindly check the 'VPLS ID' section in {node} in the selected workbook for duplicate 'VPLS ID'")
        
        return True
    
    except CustomException:
        return False

def file_vpls_checker(vpls_id,lines):
    for line in lines:
        if(line.strip().startswith(f"vpls {vpls_id}")):
            return "Exist"
    
    return "Not Exist"

def lines_separater(lines):
    required_lines = []
    for line in lines:
        if(line.startswith("vpls") and line.endswith("create")):
            # print(line)
            required_lines.append(line)
    
    return required_lines
    
# def checker(lines,dataframe,node,workbook):
    # try:
    #     dataframe.fillna("TempNA",inplace= True)
    #     thread = CustomThread(target = dataframe_checker,args = (dataframe,node))
    #     thread.daemon = True
    #     thread.start()
    #     status = thread.join()

    #     if(not status):
    #         return "Unsuccessful"
        
    #     else:
    #         thread = CustomThread(target= lines_separater, args= (lines,))
    #         thread.daemon = True
    #         thread.start()
    #         lines = thread.join()

    #         wb = load_workbook(workbook)
    #         sheetnames = wb.sheetnames

    #         if(not node in sheetnames):
    #             wb.create_sheet(node)

    #         ws = wb[node]
    #         max_row = ws.max_row

    #         if(max_row == 0):
    #             row_start = max_row+1
            
    #         if(max_row > 0):
    #             row_start = max_row+2
            
    #         col_width = []

    #         ws[f"A{row_start}"] = "VPLS Checks"
    #         ws[f"A{row_start+1}"] = "VPLS ID"
    #         ws[f"B{row_start+1}"] = "Checks"

    #         ws[f"A{row_start}"].alignment = Alignment(horizontal= 'center', vertical= 'center')
    #         ws[f"A{row_start+1}"].alignment = Alignment(horizontal= 'center', vertical= 'center')
    #         ws[f"B{row_start+1}"].alignment = Alignment(horizontal= 'center', vertical= 'center')
            
    #         ws[f"A{row_start}"].fill = PatternFill(start_color= "EF7034", end_color= "EF7034", fill_type= 'solid')
    #         ws[f"A{row_start+1}"].fill = PatternFill(start_color= "F1F10F", end_color= "F1F10F", fill_type= "solid")
    #         ws[f"B{row_start+1}"].fill = PatternFill(start_color= "F1F10F", end_color= "F1F10F", fill_type= "solid")

    #         ws[f"A{row_start}"].font = Font(bold= True)
    #         ws[f"A{row_start+1}"].font = Font(bold= True)
    #         ws[f"B{row_start+1}"].font = Font(bold= True)

    #         ws[f"A{row_start}"].border = Border(left= Side(border_style= "medium",color = "000000"), right= Side(border_style= "medium",color = "000000"), top= Side(border_style= "medium",color = "000000"), bottom= Side(border_style= "medium",color = "000000"))
    #         ws[f"A{row_start+1}"].border = Border(left= Side(border_style= "medium",color = "000000"), right= Side(border_style= "medium",color = "000000"), top= Side(border_style= "medium",color = "000000"), bottom= Side(border_style= "medium",color = "000000"))
    #         ws[f"B{row_start+1}"].border = Border(left= Side(border_style= "medium",color = "000000"), right= Side(border_style= "medium",color = "000000"), top= Side(border_style= "medium",color = "000000"), bottom= Side(border_style= "medium",color = "000000"))

    #         row_start += 2
    #         for i in range(len(dataframe)):
    #             vpls_id = dataframe.iloc[i]['VPLS ID']
    #             ws[f"A{row_start}"] = vpls_id
    #             thread = CustomThread(target= file_vpls_checker, args=(vpls_id,lines))
    #             thread.start()
    #             status = thread.join()
                
    #             ws[f"A{row_start}"].border = Border(left= Side(border_style= "medium",color = "000000"), right= Side(border_style= "medium",color = "000000"), top= Side(border_style= "medium",color = "000000"), bottom= Side(border_style= "medium",color = "000000"))
    #             ws[f"B{row_start}"].border = Border(left= Side(border_style= "medium",color = "000000"), right= Side(border_style= "medium",color = "000000"), top= Side(border_style= "medium",color = "000000"), bottom= Side(border_style= "medium",color = "000000"))

    #             match status:
    #                 case "Exist" : 
    #                     ws[f"B{row_start}"] = "Exists"
    #                     ws[f"B{row_start}"].fill = PatternFill(start = "00FF00", end = "00FF00", fill_type= "solid")
    #                     ws[f"B{row_start}"].font = Font(color = "FFFFFF")

    #                 case "Not Exist" : 
    #                     ws[f"B{row_start}"] = "Not Exists"
    #                     ws[f"B{row_start}"].fill = PatternFill(start = "", end = "", fill_type= "solid")
    #                     ws[f"B{row_start}"].font = Font(color = "FFFFFF")
                
                
    #             row_start += 1
    #         # print("after match : \n",lines,end= "\n\n")
    #         print(ws.max_column)
    #         for row_values in ws.iter_rows(values_only= True):
    #             for j,value in enumerate(row_values):
    #                 print(j)
    #                 if(len(col_width) > j):
    #                     if(col_width[j] < len(str(value))):
    #                         col_width[j] = len(str(value))
                    
    #                 else:
    #                     col_width.insert(j,len(str(value)))
            
    #         for i,col_wid in enumerate(col_width,1):
    #             ws.column_dimensions[get_column_letter(i)].width = col_wid+2

    #         wb.save(workbook)
    #         wb.close()
    #         del wb
    
    # except CustomException:
    #     return "Unuccessful"
    
    # except Exception as e:
    #     messagebox.showinfo("   Exception Occured!",f"{traceback.format_exc()}\n\n{e}")
    #     return "Unsuccessful"

def checker(lines,dataframe,node):
    try:
        dataframe.fillna("TempNA", inplace = True)
        thread = CustomThread(target= dataframe_checker,args = (dataframe,node))
        thread.daemon = True
        thread.start()
        status = thread.join()

        if(not status):
            return "Unsuccessful"
        
        else:
            thread = CustomThread(target= lines_separater, args=(lines,))
            thread.daemon = True
            thread.start()
            lines = thread.join()
            dictionary_for_status = {}

            for i in range(len(dataframe)):
                vpls_id = dataframe.iloc[i]['VPLS ID']
                thread = CustomThread(target= file_vpls_checker,args= (vpls_id,lines) )
                thread.daemon = True
                thread.start()
                dictionary_for_status[vpls_id] = []
                dictionary_for_status[vpls_id].append(dataframe.iloc[i]["VPLS Name"])
                dictionary_for_status[vpls_id].append(thread.join())



            return ["Successful",dictionary_for_status]
    
    except CustomException:
        return "Unsuccesful"
    
    except Exception as e:
        messagebox.showerror("  Exception Occured!",f"{traceback.format_exc()}\n\n{e}")
        return "Unsuccessful"
