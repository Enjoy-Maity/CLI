import pandas as pd
import traceback
import openpyxl
from tkinter import messagebox
import win32com.client as win32
from CustomException import CustomException

flag = ""
workbook1 = ""

def main(workbook):
	try:
		workbook1 = workbook
		workbook_reader = pd.ExcelFile(workbook)
		sheetnames = workbook_reader.sheet_names
		# print(sheetnames)
		# wb = pd.read_excel(workbook_reader)
		for i in sheetnames:
			if (len(i.split(".")) == 4):
				print(i)
			
	
	except CustomException:
		flag = "Unsuccessful"
		
	except Exception as e:
		messagebox.showerror(f"{traceback.format_exc()}\n\n{e}")
		flag = "Unsuccessful"
		#print(e)
	
	finally:
		excel = win32.Dispatch("Excel.Application")
		if(len(workbook) > 0):
			wb = excel.Workbooks.Open(workbook1)
			wb.Close()
		excel.Application.Quit()
		return flag
        
        
main(r"C:\Users\emaienj\Downloads\Nokia_Design Input_Template.xlsx")