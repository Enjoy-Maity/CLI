import os
import logging
import traceback
import pickle
import importlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tkinter import messagebox
from pathlib import Path

flag = ""
section_dictionary = {}


def text_file_writer(path: str, cli_dictionary: dict) -> None:
    """
    Creates the text file for node ips cli
    :param path: Path of the parent folder of the CLI text file
    :param cli_dictionary: dictionary containing ip_nodes as keys and cli as there values
    :return: None
    """

    text_file_path = os.path.join(path, "Nokia_cli.txt")

    ip_nodes = list(cli_dictionary.keys())

    cli = ''

    i = 0
    while i < len(ip_nodes):
        ip_node = ip_nodes[i]
        cli = f"{cli}************ Node IP :- {ip_node}*************\n{cli_dictionary[ip_node]}\n"
        i += 1

    logging.info(f"Writing into {text_file_path} cli:-\n{cli}")

    with open(text_file_path, 'w') as f:
        f.write(cli)
        f.close()
    del f


def excel_writer(path: str, cli_dictionary: dict) -> None:
    """
    Creates the Excel workbook containing the cli commands
    :param cli_dictionary: dictionary containing ip_nodes as keys and cli as there values
    :param path: Path of the parent folder of the CLI Excel workbook
    :return: None
    """

    excel_file_path = os.path.join(path, "CLI Execution.xlsx")
    if not os.path.exists(excel_file_path):
        workbook = Workbook()
        workbook.save(excel_file_path)
        del workbook

    workbook = load_workbook(filename=excel_file_path)
    sheetnames = workbook.sheetnames
    required_sheetnames = list(cli_dictionary.keys())
    logging.debug(f"Got the list of {required_sheetnames = }")

    border_side = Side(border_style='medium',
                       color='000000')

    i = 0
    while i < len(required_sheetnames):
        selected_sheetname = required_sheetnames[i]

        if selected_sheetname not in sheetnames:
            workbook.create_sheet(title=f"{selected_sheetname}")

        worksheet = workbook[str(selected_sheetname)]

        worksheet['A1'] = "Commands"

        worksheet['A1'].alignment = Alignment(horizontal='center',
                                              vertical='center')

        worksheet['A1'].font = Font(bold=True,
                                    name="Ericsson Hilda",
                                    size=12)

        worksheet['A1'].fill = PatternFill(start_color="FDDA0D",
                                           end_color="FDDA0D",
                                           fill_type='solid')

        worksheet['A1'].border = Border(top=border_side,
                                        bottom=border_side,
                                        left=border_side,
                                        right=border_side)

        cli_lines = cli_dictionary[selected_sheetname]
        cli_lines = cli_lines.split("\n")
        cli_lines = [line.strip() for line in cli_lines]
        cli_lines = cli_lines[1:]

        logging.info(f"{selected_sheetname}: - \n{'\n'.join(cli_lines)}")

        max_col_width = 0
        j = 0
        while j < len(cli_lines):
            if len(cli_lines[j]) > max_col_width:
                max_col_width = len(cli_lines[j])

            j += 1

        font = Font(name='Ericsson Hilda',
                    size=11)
        logging.info(f"Created the font object for {font = } for {selected_sheetname}")

        j = 0

        while j < len(cli_lines):
            worksheet[f'A{j + 2}'] = cli_lines[j]
            worksheet[f'A{j + 2}'].border = border_side
            worksheet[f'A{j + 2}'].font = font
            j += 1

        worksheet.column_dimensions['A'].width = max_col_width + 3

        i += 1

    sheetnames = workbook.sheetnames

    for sheetname in sheetnames:
        if sheetname not in required_sheetnames:
            del workbook[sheetname]

    workbook.save(filename=excel_file_path)
    workbook.close()
    del workbook


def section_wise_cli_preparation(dictionary: dict, ip_node: str) -> str:
    """
    Calls the corresponding CLI Section method for CLI Preparation
    :param ip_node: ip node for which the dictionary from the design input workbook
    :param dictionary: dictionary containing the data for the ip node
    :return: cli (str): cli from all the sections in ip_node sheet of the design sheet
    """
    global section_dictionary
    cli = ''

    logging.info(f"Starting the cli preparation for {ip_node}")

    list_of_sections = list(dictionary.keys())

    i = 0
    while i < len(list_of_sections):
        selected_section = list_of_sections[i]

        logging.info(f"Starting the cli preparation for {selected_section} of {ip_node}")
        dataframe = dictionary[selected_section]

        module_to_be_imported = section_dictionary[selected_section]

        section_cli = module_to_be_imported.main_func(dataframe, ip_node)
        cli = f'{cli}\n{section_cli}'

        i += 1

    return cli


def main_func() -> str:
    global flag
    global section_dictionary
    section_dictionary = {'VPLS-1': importlib.import_module("Main_application.Nokia.Nokia_Section_CLI_Preparation.VPLS_1_cli_prep"),
                          'VPLS-2': importlib.import_module("Main_application.Nokia.Nokia_Section_CLI_Preparation.VPLS_2_cli_prep")}
    cli = ''
    cli_dictionary = {}

    try:
        username = (os.popen(cmd='cmd.exe /C "echo %username%"').read()).strip()

        pickle_file = f"C:\\Users\\{username}\\AppData\\Local\\CLI_Automation\\Vendor_pickles\\NOKIA.pickle"

        with open(pickle_file, 'rb') as f:
            nokia_design_input_wkbk_dictionary = pickle.load(f, encoding="UTF-8")
            f.close()

        host_details_file_path_file = f"C:\\Users\\{username}\\AppData\\Local\\CLI_Automation\\host_details_file_path.txt"

        with open(host_details_file_path_file, 'r') as f:
            parent_folder_of_selected_host_details_wbk = os.path.dirname(f.readline())
            f.close()

        ip_nodes = list(nokia_design_input_wkbk_dictionary.keys())

        i = 0
        while i < len(ip_nodes):
            cli = ''
            logging.debug(f"Passing the arguments to the section_wise_cli_preparation function for node :{ip_nodes[i]}")

            ip_node_dictionary = nokia_design_input_wkbk_dictionary[ip_nodes[i]]

            cli = section_wise_cli_preparation(dictionary=ip_node_dictionary,
                                               ip_node=ip_nodes[i])

            logging.info(f"Got the Cli for \'{ip_nodes[i]}\' ")

            cli = f"{cli}admin save"
            cli_dictionary[ip_nodes[i]] = cli
            i += 1

        parent_folder_for_cli_preparation_wkbk = os.path.join(parent_folder_of_selected_host_details_wbk, "CLI")
        parent_folder_for_cli_preparation_wkbk = os.path.join(parent_folder_for_cli_preparation_wkbk, "Nokia")

        Path(parent_folder_for_cli_preparation_wkbk).mkdir(exist_ok=True, parents=True)

        text_file_writer(parent_folder_for_cli_preparation_wkbk, cli_dictionary)

        excel_writer(parent_folder_for_cli_preparation_wkbk, cli_dictionary)

        if flag != 'Unsuccessful':
            flag = 'Successful'

    except Exception as e:
        logging.error(f"{traceback.format_exc()}\n\tException Occurred!===>\nTitle==>{type(e)}\n\t\tMessage==>{str(e)}")
        flag = 'Unsuccessful'
        messagebox.showerror(title="Exception Occurred!",
                             message=str(e))

    finally:
        logging.info(f"Returning flag => {flag}")
        logging.shutdown()

        return flag
