import traceback
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
from openpyxl.utils import get_column_letter

def vpls_checks_dumper(vpls_section_checks_dictionary,node,workbook):
	try:
		wb = load_workbook(workbook)
		ws = wb[node]

		max_row = ws.max_row
		start_row = 0

		if(max_row == 0):
			start_row = 0
		
		else:
			start_row = max_row + 2
		
		ws[f"A{start_row}"] = "VPLS"
		ws[f"A{start_row}"].font = Font(color="000000", bold = True)
		ws[f"A{start_row}"].alignment = Alignment(horizontal= "center",vertical= "center")
		ws[f"A{start_row}"].fill = PatternFill(start_color = "FFFF00",end_color = "FFFF00",fill_type = "solid")
		ws[f"A{start_row}"].border = Border(left  = Side(border_style= "medium",color = "000000"), 
				      						right = Side(border_style= "medium",color = "000000"),
											top = Side(border_style= "medium",color = "000000"),
											bottom = Side(border_style= "medium",color = "000000"))

		ws[f"A{start_row+1}"] = "VPLS ID"
		ws[f"A{start_row+1}"].font = Font(color="000000",bold = True)
		ws[f"A{start_row+1}"].alignment = Alignment(horizontal= "center",vertical= "center")
		ws[f"A{start_row+1}"].fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
		ws[f"A{start_row+1}"].border = Border(left  = Side(border_style= "medium",color = "000000"),
				      						right = Side(border_style= "medium",color = "000000"),
											top = Side(border_style= "medium",color = "000000"),
											bottom = Side(border_style= "medium",color = "000000"))
		
		ws[f"B{start_row+1}"] = "VPLS Name"
		ws[f"B{start_row+1}"].font = Font(color="000000",bold = True)
		ws[f"B{start_row+1}"].alignment = Alignment(horizontal= "center",vertical= "center")
		ws[f"B{start_row+1}"].fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
		ws[f"B{start_row+1}"].border = Border(left  = Side(border_style= "medium",color = "000000"),
				      						right = Side(border_style= "medium",color = "000000"),
											top = Side(border_style= "medium",color = "000000"),
											bottom = Side(border_style= "medium",color = "000000"))
		
		ws[f"C{start_row+1}"] = "VPLS Checks"
		ws[f"C{start_row+1}"].font = Font(color="000000",bold = True)
		ws[f"C{start_row+1}"].alignment = Alignment(horizontal= "center",vertical= "center")
		ws[f"C{start_row+1}"].fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
		ws[f"C{start_row+1}"].border = Border(left  = Side(border_style= "medium",color = "000000"),
				      						right = Side(border_style= "medium",color = "000000"),
											top = Side(border_style= "medium",color = "000000"),
											bottom = Side(border_style= "medium",color = "000000"))
		
		row = start_row + 2
	
		# neo_vpls_section_checks_dictionary = {key:value for key,value in vpls_section_checks_dictionary.items() if value[1] == "Not Exist"}
		# print(neo_vpls_section_checks_dictionary)

		# neo_vpls_section_checks_dictionary = {key:value for key,value in vpls_section_checks_dictionary.items() if value[1] == "Exist"}
		# print(neo_vpls_section_checks_dictionary)
		neo_vpls_section_checks_dictionary = {}

		for key,value in vpls_section_checks_dictionary.items():
			if(value[1] == "Not Exist"):
				neo_vpls_section_checks_dictionary[key] = value
		
		for key,value in vpls_section_checks_dictionary.items():
			if(value[1] == "Exist"):
				neo_vpls_section_checks_dictionary[key] = value
		
		# print("\n\nneo_vpls_section_checks\n\n",neo_vpls_section_checks_dictionary)

		vpls_section_checks_dictionary = neo_vpls_section_checks_dictionary
		
		del neo_vpls_section_checks_dictionary
		
		for vpls_id,value in vpls_section_checks_dictionary.items():
			ws[f"A{row}"] = vpls_id
			ws[f"B{row}"] = value[0]
			ws[f"C{row}"] = value[1]
			
			
			ws[f"A{row}"].border = Border(left  = Side(border_style= "medium",color = "000000"), 
				      					  right = Side(border_style= "medium",color = "000000"),
										  top = Side(border_style= "medium",color = "000000"),
										  bottom = Side(border_style= "medium",color = "000000"))
			
			ws[f"B{row}"].border = Border(left  = Side(border_style= "medium",color = "000000"),
				      					  right = Side(border_style= "medium",color = "000000"),
										  top = Side(border_style= "medium",color = "000000"),
										  bottom = Side(border_style= "medium",color = "000000"))
			ws[f"C{row}"].border = Border(left  = Side(border_style= "medium",color = "000000"),
				      					  right = Side(border_style= "medium",color = "000000"),
										  top = Side(border_style= "medium",color = "000000"),
										  bottom = Side(border_style= "medium",color = "000000"))
			match value[1]:
				case "Exist" :
					ws[f"C{row}"].fill = PatternFill(start_color = "00FF00", end_color = "00FF00", fill_type = "solid")
					ws[f"C{row}"].font = Font(color = "000000")
				
				case "Not Exist":
					ws[f"C{row}"].fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
					ws[f"C{row}"].font = Font(color = "000000")
			
			row += 1
		
		col_width = []
		for row_values in ws.iter_rows(values_only= True):
			for j,value in enumerate(row_values):
				# print(j)
				if(len(col_width) > j):
					if(col_width[j] < len(str(value))):
						col_width[j] = len(str(value))
				
				else:
					col_width.insert(j,len(str(value)))
		
		for i,col_wid in enumerate(col_width,1):
			ws.column_dimensions[get_column_letter(i)].width = col_wid+2
		
		wb.save(workbook)
		wb.close()
		del wb

		return "Successful "
	
	except Exception as e:
		messagebox.showerror("	Exception Occured!",f"{traceback.format_exc()}\n\n{e}")
		return "Unsuccessful"