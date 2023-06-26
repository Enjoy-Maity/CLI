import pandas as pd							# for dataframe manipulation
import traceback							# for error traceback to line number and statement causing error
import pythoncom							# to handle windows processes
from openpyxl import load_workbook,Workbook	# to handle the excel files.
from tkinter import messagebox				# to create message dialogues
import win32com.client as win32				# for MS Office application handling
from CustomException import CustomException	# Custom class for handling custom raised exceptions
from Custom_Thread import CustomThread		# Custom class Threads with return value
from pathlib import Path					# for handling paths of files and folders
import os
# import numpy as np
# import re

flag = ""
workbook1 = ""
workbook2 = ""

def section_splitter(dataframe):
	string_to_be_found = "#Secti0n_MPBN"
	dataframe = dataframe.fillna("TempNA")
	section_dictionary = {}
	
	for i in range(0,len(dataframe)):
		if(dataframe.iloc[i][0].__contains__(string_to_be_found)):
			section = dataframe.iloc[i][0].split(string_to_be_found)[0]
			unrefined_columns = dataframe.loc[i+1].tolist()
			columns = []
			
			for column_header in unrefined_columns:
				if(column_header == "TempNA"):
					break
				else:
					columns.append(column_header)

			j = i+2
			dictionary_for_columns = {}
			for k in columns:
				dictionary_for_columns[k] = []
			
			while((j)<len(dataframe)):
				if(((j+2) < len(dataframe)) and (dataframe.iloc[j+2][0].__contains__(string_to_be_found))):
					break
				else:
					
					for k in range(len(columns)):
						dictionary_for_columns[columns[k]].append(dataframe.iloc[j][k])
					j+=1
			i = j+2
			
			section_dictionary[section] = pd.DataFrame(dictionary_for_columns,columns=columns)
			copy_df = section_dictionary[section].copy()
			section_dictionary[section].drop_duplicates(keep = "first", inplace = True)
			if(section.strip().upper() == "VPLS"):
				if(len(copy_df)> len(section_dictionary[section])):
					duplicate_vpls_id_list = []
					duplicate_value_counts = copy_df.duplicated(keep='first').tolist()
					# print(duplicate_value_counts)
					# print({i:j for i,j in enumerate(duplicate_value_counts)})
					# print(copy_df['VPLS ID'].value_counts())
					
					
					for indx in range(0,len(duplicate_value_counts)):
						if(duplicate_value_counts[indx]):
							duplicate_vpls_id_list.append(copy_df.iloc[indx]['VPLS ID'])

					messagebox.showinfo("Duplicate VPLS ID Rows found",f"Duplicate rows found for below vpls ids:\n{', '.join(str(i) for i in duplicate_vpls_id_list)}")

			section_dictionary[section].replace("TempNA","",inplace = True)
			
		else:
			continue
	
	return section_dictionary

def main(workbook):
	try:
		global workbook1;
		global workbook2;
		workbook1 = workbook
		global flag;
		workbook_reader = pd.ExcelFile(workbook)
		sheetnames = workbook_reader.sheet_names
		
		nodes = []
		for i in sheetnames:
			if (len(i.split(".")) == 4):
				nodes.append(i)
		
		if(len(nodes) == 0):
			raise CustomException('	Node details absent!',f"The workbook {workbook1} doesn't contain the Node details, Please Check!")
		
		if(len(nodes) < len(sheetnames)):
			raise CustomException('	Node details incorrect!',f"The workbook {workbook1} contain incorrect Node details, Please Check!")
		
		else:
			cli = ""
			dictionary_for_node_section = {}
			main_folder = workbook.split("/")
	
			folder_for_cli = f"{'/'.join(main_folder[:-1])}/CLI"

			Path(folder_for_cli).mkdir(parents=True,exist_ok=True)
			
			file_to_write = f"{folder_for_cli}/Cli.xlsx"

			
			
			for sheetname in nodes:
				df = pd.read_excel(workbook_reader,sheetname)
				dictionary_for_node_section[sheetname] = section_splitter(df)
		
			keys = dictionary_for_node_section.keys()
			temp_flag = "Unsuccessful"

			for key in keys:
				cli = ""
				secondary_keys = dictionary_for_node_section[key].keys()
				for second_key in secondary_keys:
					if((second_key.upper() == "VPLS") and (len(dictionary_for_node_section[key][second_key]) > 0)):

						unique_vpls_id_len = len(dictionary_for_node_section[key][second_key]['VPLS ID'].unique())
						dataframe_len = len(dictionary_for_node_section[key][second_key])

						if(unique_vpls_id_len < dataframe_len):
							duplicate_vpls_id_list = []
							duplicate_value_counts = dictionary_for_node_section[key][second_key]['VPLS ID'].value_counts().index.tolist()
							
							for vpls_id in duplicate_value_counts:
								if(int(dictionary_for_node_section[key][second_key]['VPLS ID'].value_counts()[vpls_id]) > 1):
									duplicate_vpls_id_list.append(vpls_id)
							
							raise CustomException("	Duplicate VPLS ID found!",f"Kindly Check the 'VPLS' Section for below duplicate 'VPLS ID':\n{', '.join(str(vpls_id_num) for vpls_id_num in duplicate_vpls_id_list)}")
						
						else:
							import VPLS_CLI_prepare
							# print(key)
							cli = f"{cli}configure service\n"
							thread = CustomThread(target = VPLS_CLI_prepare.cli_maker, args=(dictionary_for_node_section[key][second_key],key,cli))
							thread.daemon = True
							thread.start()
							cli = thread.join()

							if(cli.strip() == "Unsuccessful"):
								temp_flag = "Unsuccessful"
								raise CustomException("	Exception Occured!",f"Cli Creation is Unsuccessful")
							
							
							cli = f"{cli}exit all\n"
			
				cli = f"{cli}\
						admin save"
				
				if(temp_flag != "Unsuccesful"):
					# if(Path(file_to_write).exists()):
					# 	os.remove(file_to_write)
					# 	wb = Workbook()
					# 	wb.save(file_to_write)
					# 	wb.close()
					# 	del wb

					import cli_dumper
					cli_dumper.cli_writer(cli,file_to_write,key)

				
				workbook2 = file_to_write
				wb = load_workbook(file_to_write)
				sheetnames = wb.sheetnames
				# print(sheetnames)

				if(len(sheetnames) > 1):
					for sheetname in sheetnames:
						if(not sheetname in nodes):
							del wb[sheetname]
				
				wb.save(file_to_write)
				wb.close()

			flag = "Successful"
			messagebox.showinfo("	CLI Created Successfully!",f"CLI created at {file_to_write} successfully!")
	
	except CustomException:
		flag = "Unsuccessful"
	
	except pythoncom.com_error as ex:
		# Convert the error to a human-readable form
		error_code, error_msg, error_desc, error_help_file, error_help_context, error_sc_info = ex.args
		messagebox.showerror("	Exception Occured!",f"{traceback.format_exc()}\n\n{e}\n\nDescription: {error_desc}")
		flag = "Unsuccessful"

	except Exception as e:
		messagebox.showerror("	Exception Occured!",f"{traceback.format_exc()}\n\n{e}")
		flag = "Unsuccessful"
		
	
	finally:
		# global workbook1;

		# excel = win32.Dispatch("Excel.Application")

		# if((len(workbook1) > 0) and (Path(workbook1).exists())):
		# 	wb = excel.Workbooks.Open(workbook1)
		# 	wb.Close()
		
		# if((len(workbook2) > 0) and (Path(workbook2).exists())):
		# 	wb = excel.Workbooks.Open(workbook2)
		# 	wb.Close()
		# excel.Application.Quit()
		return flag
        
        
main(r"C:/Users/emaienj/Downloads/Nokia_Design Input_Template.xlsx")