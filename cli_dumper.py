from pathlib import Path
import win32com.client as win32
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side


def cli_writer(cli,workbook,node):
	
	lines = cli.splitlines()
	if(Path(workbook).exists() == False):
		wb = Workbook()
		wb.save(workbook)
		wb.close()

	# excel_file = win32.Dispatch("Excel.Application")
	# wb = excel_file.Workbooks.Open(workbook)
	# sheet_exists = False
	# for sheet in wb.Sheets:
	# 	if(sheet.Name == node):
	# 		sheet_exists = True
	
	# if(not sheet_exists):
	# 	ws = wb.Worksheets.Add()
	# 	ws.Name = node
	
	# else:
	# 	ws = wb.Worksheets(node)
	# 	lastrow = sheet.UsedRange.Rows.Count		#to get the last row number	

	# 	for i in range(1,lastrow+1):
	# 		ws.Cells(i,'A').Delete()

	# # lastcol = sheet.UsedRange.Columns.Count	#to get the last column number
	

	# ws.Cells(1,1).Value = "Command"
	# # ws.Cells(1,1).Interior.Color = "#FDDA0D"
	# ws.Cells(1,1).Borders.LineStyle = 1
	# ws.Cells(1,1).Borders.Weight = 2
	
	# i=1
	# for line in lines:
	# 	i+=1
	# 	ws.Cells(i,'A').Value = line.strip()
	# 	ws.Cells(i,1).Borders.LineStyle = 1
	# 	ws.Cells(i,1).Borders.Weight = 2

	# wb.SaveAs(workbook)
	# wb.Close()

	# excel_file.Application.Quit()
	
	wb = load_workbook(workbook)
	sheetnames = wb.sheetnames
	
	if(not node in sheetnames):
		wb.create_sheet(node)
	
	ws = wb[node]
	
	max_rows_occupied = ws.max_row

	if(max_rows_occupied > 1):
		for i in range(1,max_rows_occupied + 1):
			# ws[f'A{i}'] = None
			del ws[f'A{i}']

	max_col_width = 0

	for line in lines:
		if(max_col_width < len(line.strip())):
			max_col_width = len(line.strip())
	
	ws.column_dimensions['A'].width = max_col_width + 2

	ws['A1'] = "Commands"
	ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
	ws['A1'].font = Font(bold=True)
	ws['A1'].fill = PatternFill(start_color= "FDDA0D",end_color= "FDDA0D",fill_type= "solid")
	ws['A1'].border = Border(top = Side(border_style = 'medium',color = '000000'),bottom = Side(border_style = 'medium',color = '000000'),left = Side(border_style = 'medium',color = '000000'),right = Side(border_style = 'medium',color = '000000'))

	# print(f"last cli line : {lines[-1]}")
	i = 2
	for line in lines:
		ws[f'A{i}'] = line.strip()
		ws[f'A{i}'].border = Border(top = Side(border_style = 'medium',color = '000000'),bottom = Side(border_style = 'medium',color = '000000'),left = Side(border_style = 'medium',color = '000000'),right = Side(border_style = 'medium',color = '000000'))
		i+=1

	wb.save(workbook)
	wb.close()