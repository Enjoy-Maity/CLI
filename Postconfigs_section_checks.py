import traceback
import pythoncom
import os
# import shutil
import pandas as pd
from Checks_Writer import main_driver_func
# import win32com.client as win32
from tkinter import messagebox
from pathlib import Path
from CustomException import CustomException
from openpyxl import Workbook,load_workbook
from Custom_Thread import CustomThread

workbook1 = ""
workbook2 = ""
flag = ""

def line_stripper(lines):
	neo_lines = []
	for line in lines:
		neo_lines.append(line.strip())

	return neo_lines

def section_splitter(dataframe):
	string_to_be_found = "#Secti0n_MPBN"
	dataframe = dataframe.fillna("TempNA")
	section_dictionary = {}
	
	for i in range(0,len(dataframe)):
		if(dataframe.iloc[i][0].__contains__(string_to_be_found)):
			section = dataframe.iloc[i][0].split(string_to_be_found)[0]
			unrefined_columns = dataframe.loc[i+1].tolist()
			columns = []
			
			for column_header in unrefined_columns:
				if(column_header == "TempNA"):
					break
				else:
					columns.append(column_header)

			j = i+2
			dictionary_for_columns = {}
			for k in columns:
				dictionary_for_columns[k] = []
			
			while((j)<len(dataframe)):
				if(((j+2) < len(dataframe)) and (dataframe.iloc[j+2][0].__contains__(string_to_be_found))):
					break
				else:
					for k in range(len(columns)):
						dictionary_for_columns[columns[k]].append(dataframe.iloc[j][k])
					j+=1
			i = j+2
			
			section_dictionary[section] = pd.DataFrame(dictionary_for_columns,columns=columns)
			section_dictionary[section].drop_duplicates(keep = False,inplace = True)
			section_dictionary[section].replace("TempNA","",inplace = True)
			
		else:
			continue
	
	return section_dictionary

def main(workbook):
	try:
		global workbook1;
		global workbook2;
		workbook1 = workbook
		global flag;
		workbook_reader = pd.ExcelFile(workbook)
		sheetnames = workbook_reader.sheet_names
		
		nodes = []
		for i in sheetnames:
			if (len(i.split(".")) == 4):
				nodes.append(i)
		
		if(len(nodes) == 0):
			raise CustomException('	Node details absent!',f"The workbook {workbook1} doesn't contain the Node details")
		
		if(len(nodes) > 0):
			parent_folder = os.path.dirname(workbook)
			files_inside_the_parent_folder = os.listdir(parent_folder)
			
			for node in nodes:
				if(not node in files_inside_the_parent_folder):
					raise CustomException("	Post Config File Absent!",f"Post Config File for {node} absent in {parent_folder}, Kindly Check!")
		
			global workbook2;
			dictionary_for_node_section = {}
			folder_for_checks = os.path.join(parent_folder,"Checks")
			Path(folder_for_checks).mkdir(exist_ok=True,parents=True)

			file_for_checks = os.path.join(folder_for_checks,"Checks.xlsx")

			dictionary_for_node_section_checks = {}

			for node in nodes:
				df = pd.read_excel(workbook,node)
				dictionary_for_node_section[node] = section_splitter(df)
				dictionary_for_node_section_checks[node] = {}

			

			for node in nodes:
				secondary_keys = dictionary_for_node_section[node].keys()

				file_for_postconfig = os.path.join(parent_folder,node)
				# print(node)
				with open(file_for_postconfig,mode = "r",encoding = "utf-8") as f:
						lines = f.readlines()
				
				thread = CustomThread(target= line_stripper,args=(lines,))
				thread.daemon = True
				thread.start()
				neo_lines = thread.join()
				temp_flag = 0

				for key in secondary_keys:
					if((key.strip() == "VPLS") and (len(dictionary_for_node_section[node][key]) > 0)):
						unique_vpls_id = dictionary_for_node_section[node][key]['VPLS ID'].dropna().unique()
						
						if(len(unique_vpls_id) < len(dictionary_for_node_section[node][key])):
							raise CustomException("	Blank VPLS_ID found",f"Kindly check the 'VPLS ID' in 'VPLS' Section in {node} sheet for blank fields")
						
						for vpls_id in unique_vpls_id:
							if(not str(vpls_id).isnumeric()):
								raise CustomException("	Non-Integer value found!",f"Kindly check the 'VPLS ID' in 'VPLS' Section in {node} sheet for non-integer fields")
							
						import VPLS_section_checker
						# print(dictionary_for_node_section[node][key])
						# print(node)
						return_value = VPLS_section_checker.checker(neo_lines,dictionary_for_node_section[node][key],node)
						# print("\n\n")
						# thread.daemon = True
						# thread.start()
						
						# print(return_value)
						
						if(isinstance(return_value,list)):
							dictionary_for_node_section_checks[node][key] = return_value[1]
						else:
							temp_flag = 1
							break
				
				
			if(temp_flag == 0):
				if(Path(file_for_checks).exists()):
					# import time
					# shutil.rmtree(os.path.dirname(file_for_checks))
					os.remove(file_for_checks)
					
				
				wb = Workbook(write_only= True)
				wb.create_sheet(title="Sheet 0",index = 0)
				wb.save(file_for_checks)
				wb.close()
				del wb

				workbook2 = file_for_checks
					

				# thread = CustomThread(target = main_driver_func, args = (dictionary_for_node_section_checks,file_for_checks))
				# thread.daemon = True
				# thread.start()
				# flag = thread.join()
				flag = main_driver_func(dictionary_for_node_section_checks,file_for_checks)
			
			else:
				flag = "Unsuccessful"

			if(flag.strip() == "Successful"):
				wb = load_workbook(file_for_checks)
				sheets = wb.sheetnames
				for sheet in sheets:
					print(sheet)
					if(not sheet in nodes):
						del wb[sheet]
				wb.save(file_for_checks)
				wb.close()
				del wb

				messagebox.showinfo("	Checks workbook successfully created!",f"Workbook for Checks created at path: {file_for_checks}")

	except CustomException:
		flag = "Unsuccessful"
	
	except pythoncom.com_error as ex:
		# Convert the error to a human-readable form
		error_code, error_msg, error_desc, error_help_file, error_help_context, error_sc_info = ex.args
		messagebox.showerror("	Exception Occured!",f"{traceback.format_exc()}\n\n{e}\n\nDescription: {error_desc}")
		flag = "Unsuccessful"

	except Exception as e:
		messagebox.showerror("	Exception Occured!",f"{traceback.format_exc()}\n\n{e}")
		flag = "Unsuccessful"
		
	
	finally:
		# global workbook2;

		# excel = win32.Dispatch("Excel.Application")
		# excel.Visible = False

		# if((len(workbook1) > 0) and (Path(workbook1).exists())):
		# 	wb = excel.Workbooks.Open(workbook1)
		# 	wb.Close()
		
		
		# if((len(workbook2) > 0) and (Path(workbook2).exists())):
		# 	wb = excel.Workbooks.Open(workbook2)
		# 	wb.Close()
		
		# excel.Application.Quit()
		return flag
	
# main(r"C:\Users\emaienj\Downloads\Nokia_Design Input_Template.xlsx")