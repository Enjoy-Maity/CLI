import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
from openpyxl.utils import get_column_letter
from tkinter import messagebox
# from Custom_Thread import CustomThread
from CustomException import CustomException

def main_driver_func(dictionary,workbook):
    try:
        node_keys = dictionary.keys()
        status = "Successful"
        
        wb = load_workbook(workbook)
        sheetnames = wb.sheetnames
        
        for node in node_keys:
            if(not node in sheetnames):
                wb.create_sheet(title = node)
        wb.save(workbook)
        wb.close()
        del wb
        
        for node in node_keys:
            sections = dictionary[node].keys()
            for section in sections:
                if ((section.strip().upper() == "VPLS") and (status.strip() == "Successful")):
                    import vpls_section_checks_writer
                    # dataframe = dictionary[node][section]
                    status = vpls_section_checks_writer.vpls_checks_dumper(dictionary[node][section],node, workbook)
                    
        
        return status
        
    except CustomException:
        return "Unsuccessful"
	
    except Exception as e:
        messagebox.showerror("	Exception Occurred!",f"{traceback.format_exc()}\n\n{e}")
        return "Unsuccessful"